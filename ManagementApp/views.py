from ClassroomsPlatform.models import Classroom,ClassroomPresence,ClassroomAverage
from AssignmentPlatform.models import Assignment,AssignmentScore,AssignmentAverage
from Frontend import scripts
from Frontend import sms_manager
from Frontend.upload_manager import auto_upload,process_content_urls
from ExamsPlatform.models import Question,Exam,ExamScore,ExamScoreOffline,ExamAverage
from StudentsInfo.models import DirectMoney,SignedCheck,StudentUser,Books,Notification, UserNotification,StudentYearRecord,YearExamRecord,YearAssignmentRecord,AttendanceRecord
from StudentsInfo.serializers import NotificationSerializer, UserNotificationSerializer
from django.contrib.auth.models import Group,User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.db import transaction
from knox.models import AuthToken
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter, inline_serializer
from drf_spectacular.types import OpenApiTypes
from rest_framework import serializers as rfs
from rest_framework.decorators import api_view,action
from rest_framework.response import Response
from rest_framework import status,viewsets,views,parsers
# from rest_framework.parsers import MultiPartParser
from rest_framework.decorators import api_view,permission_classes
from rest_framework.permissions import IsAuthenticated,IsAdminUser,BasePermission,AllowAny
from rest_framework.throttling import ScopedRateThrottle, UserRateThrottle
from .serializers import GroupsSerializer,GroupSerializer,ClassroomSerializer,ClassroomsSerializer,AssignmentSerializer,ExamSerializer,QuestionSerializer,SmallUserSerializer,UserSerializer,AssignmentScoreSerializer,AssignmentScoresSerializer,SignedCheckSerializer,DirectMoneySerializer,ExamScoresSerializer,ExamSerializerWithQuestion,CreateStudentUserSerializer,UpdateStudentUserSerializer,ExamSerializerWithGroup,AssignmentSerializerWithGroup,BooksSerializer,BooksUserSerializer,ExamScoresSerializerWithExamName,ExamAverageNCSerializer,StudentYearRecordSerializer,AttendanceRecordSerializer
import pandas as pd
import io
import xlsxwriter
from django.http import HttpResponse
import openpyxl
import jdatetime
from datetime import date,datetime
from django.shortcuts import get_object_or_404
import logging
from django.utils import timezone


class IsStaffUser(BasePermission):
    def has_permission(self,request,view):
        return request.user and request.user.is_staff

@extend_schema(
    request=inline_serializer(name='AdminChangePasswordRequest', fields={
        'username': rfs.CharField(),
        'new_password': rfs.CharField(),
    }),
    responses=OpenApiTypes.OBJECT)
@api_view(['POST'])
@permission_classes([IsAdminUser])
def change_user_password(request):
    username = request.data.get('username')
    new_password = request.data.get('new_password')

    if not username or not new_password:
        return Response({'error': 'Please provide both username and new_password'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    # ضدِ ارتقای دسترسی: فقط superuser می‌تواند رمز حساب‌های staff/superuser را عوض کند
    if (user.is_staff or user.is_superuser) and not request.user.is_superuser:
        return Response({'error': 'Only a superuser can change a staff/superuser password'},
                        status=status.HTTP_403_FORBIDDEN)

    # الزامِ رمزِ قوی فقط برای حساب‌های مدیریتی (کدملیِ دانش‌آموزان دست‌نخورده می‌ماند)
    if user.is_staff or user.is_superuser:
        try:
            validate_password(new_password, user=user)
        except ValidationError as e:
            return Response({'error': 'weak password', 'messages': e.messages},
                            status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)
    user.save()
    # باطل‌کردن توکن‌های فعال کاربر هدف تا با رمز جدید دوباره لاگین کند
    AuthToken.objects.filter(user=user).delete()

    return Response({'success': 'Password updated successfully'}, status=status.HTTP_200_OK)


# بایگانی پایان سال از HTTP خارج شد و به management command منتقل شد (امنیت + بکاپ اجباری):
#     python manage.py archive_year
# منطق در ManagementApp/services.py::archive_academic_year

# @api_view(['GET'])
# @permission_classes([IsAdminUser])
# def zappier(request):
#     scripts.temporaryscript()
#     return Response({"message": "Run Successfully"}, status=status.HTTP_200_OK)    

# @api_view(['GET'])
# @permission_classes([IsAdminUser])
# def zappier1(request):
#     scripts.cfas()
#     return Response({"message": "Run Successfully"}, status=status.HTTP_200_OK)    



class SmsRateThrottle(UserRateThrottle):
    """نرخِ اختصاصیِ ارسال پیامک (اسکوپ 'sms' = ۳۰ در ساعت).

    چون هر پیامک هزینه‌ی مالی دارد، اندپوینت‌های ارسال پیامک نباید زیرِ نرخِ
    عمومیِ کاربر (۶۰۰ در دقیقه) قرار بگیرند. کلید بر اساس کاربرِ لاگین‌کرده است.
    """
    scope = 'sms'


#########################   Users - Groups - Attendance - SMS
class GroupsIndex (viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]
    search_fields=('name','group_time','group_day','group_grade','group_gender')
    def get_serializer_class(self):
        if self.action == 'list':
            return GroupsSerializer
        return GroupSerializer

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser])
    def attendance(self, request, pk=None):
        try:
            # دریافت تکلیف
            # assignment = get_object_or_404(Assignment, assignment_id=pk)
            pass
        except Exception as e:
            return Response(
                {"message": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser])
    def export_exam_scores(self, request, pk):
        group_id = pk
        group = Group.objects.get(id=group_id)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=exam_scores.xlsx'

        wb = openpyxl.Workbook()
        exams = Exam.objects.filter(exam_group_id=group_id, exam_finished=True)
        for exam in exams:
            ws = wb.create_sheet(exam.ExamName)
            ws.append(['Student', 'Score', 'Exam Presence', 'Questions Answer List', 'User Choice', 'Wrong Counts', 'None Counts'])
            exam_scores = ExamScore.objects.filter(exam=exam)
            for exam_score in exam_scores:
                student = exam_score.exam_average_reffer.user
                student_name = f'{student.first_name} {student.last_name}'  # or student.username
                if exam_score.exam_peresence:
                    score = exam_score.score
                    presence = exam_score.exam_peresence
                    questions_answer_list = exam_score.questions_answer_list
                    user_choice = exam_score.user_choice
                    wrong_counts = exam_score.wrong_counts
                    none_counts = exam_score.none_counts
                else:
                    score = '0'#absent
                    presence = 'غایب'
                    questions_answer_list = ''
                    user_choice = ''
                    wrong_counts = ''
                    none_counts = ''
                ws.append([student_name, score, presence, questions_answer_list, user_choice, wrong_counts, none_counts])
        wb.save(response)
        return response    
        
    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser])
    def export_assignment_scores(self, request, pk):
        group_id = pk
        # group = Group.objects.get(id=group_id)
        # # 
        # filename = f'{group.name}_Assignment_summary.xlsx'
        # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # # response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = f'attachment; filename={filename}'
        # # 
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] ='attachment; filename=assignment_scores.xlsx'
        # group_id = pk
        # group = Group.objects.get(id=group_id)
        # filename = re.sub(r'[<>:"/\\|?*]', '', f'{group.name}_Assignment_summary.xlsx')  # remove invalid characters
        # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = f'attachment; filename={filename}'
        wb = openpyxl.Workbook()
        assignments = Assignment.objects.filter(assignment_group_id=group_id, assignment_finished=True)
        for assignment in assignments:
            ws = wb.create_sheet(assignment.AssignmentName)
            ws.append(['Student', 'Score', 'Assignment Presence', 'Assignment Marked'])
            assignment_scores = AssignmentScore.objects.filter(assignment=assignment)
            for assignment_score in assignment_scores:
                student = assignment_score.assignment_average_reffer.user
                student_name = f'{student.first_name} {student.last_name}'  # or student.username
                if assignment_score.assignment_presence:
                    score = assignment_score.score
                    presence = assignment_score.assignment_presence
                    marked = assignment_score.assignment_marked
                else:
                    score = '0'
                    presence = 'غایب'
                    marked = ''
                ws.append([student_name, score, presence, marked])
        wb.save(response)
        return response
    
class UsersIndex (viewsets.ModelViewSet):
    queryset = User.objects.filter(is_staff=False).order_by('-date_joined')
    serializer_class = UserSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]
    search_fields=('first_name','last_name','username','groups__name','studentuser__student_type','groups__group_grade','groups__group_gender','studentuser__student_school','studentuser__mother_number','studentuser__registration_date')

    def get_queryset(self):
        # رفع N+1: برای detail سریالایزر تودرتوی سنگین prefetch می‌شود، برای list سبک‌تر
        qs = User.objects.filter(is_staff=False).order_by('-date_joined')
        if self.action == 'retrieve':
            qs = qs.select_related('studentuser', 'examaverage', 'assignmentaverage').prefetch_related(
                'groups',
                'examaverage__examscore_set__exam',
                'assignmentaverage__assignmentscore_set__assignment',
                'directmoney_set', 'signedcheck_set',
            )
        elif self.action == 'list':
            qs = qs.select_related('studentuser').prefetch_related('groups')
        return qs

    # def list(self, request):
    #     ordering = request.query_params.get('ordering', '-date_joined')
    #     queryset = User.objects.filter(is_staff=False).order_by(ordering)
    #     serializer = SmallUserSerializer(queryset, many=True)
    #     return Response(serializer.data)
    
    def get_serializer_class(self):
        if self.action == 'list':
            return SmallUserSerializer
        return UserSerializer

    def create(self, request, *args, **kwargs):
        user_data = request.data.copy()
        password = user_data.pop('password', None)
        user_serializer = self.get_serializer(data=user_data)
        user_serializer.is_valid(raise_exception=True)
        user_instance = user_serializer.save()
        if password:
            user_instance.set_password(password)
            user_instance.save()
    
        student_user_data = request.data.get('studentuser')
        if student_user_data:
            student_user_data['student_user'] = user_instance.id
            student_user_serializer = CreateStudentUserSerializer(data=student_user_data)
            student_user_serializer.is_valid(raise_exception=True)
            student_user_serializer.save()
    
        headers = self.get_success_headers(user_serializer.data)
        return Response(user_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def perform_create(self, serializer):
        instance = serializer.save()
        student_user = instance.studentuser
        student_user.student_group_info()
        student_user.create_averages()

    def partial_update(self, request, *args, **kwargs):
        # Update User instance
        user_instance = self.get_object()
        user_serializer = self.get_serializer(user_instance, data=request.data, partial=True)
        user_serializer.is_valid(raise_exception=True)
        user_serializer.save()
    
        # Update StudentUser instance
        student_user_data = request.data.get('studentuser')
        if student_user_data:
            student_user_instance = user_instance.studentuser
            student_user_serializer = UpdateStudentUserSerializer(student_user_instance, data=student_user_data, partial=True)
            student_user_serializer.is_valid(raise_exception=True)
            student_user_serializer.save()
            student_user_instance.student_group_info()
            student_user_instance.create_averages()
        return Response(user_serializer.data)
        
    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser])
    def toggle_activation(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        return Response({'status': 'success', 'is_active': user.is_active},status=status.HTTP_200_OK)
        
    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser])
    def is_active(self, request, pk=None):
        user = self.get_object()
        return Response({'is_active': user.is_active},status=status.HTTP_200_OK)  
    
    @action(detail=True, methods=['get', 'patch'], permission_classes=[IsAdminUser])
    def user_books(self, request, pk=None):
        user = self.get_object()
        if request.method == 'GET':
            books = user.books_set.all()
            serializer = BooksUserSerializer(books, many=True)
            return Response(serializer.data,status=status.HTTP_200_OK)
        elif request.method == 'PATCH':
            book_ids = request.data.get('book_ids', [])
            books = Books.objects.filter(id__in=book_ids)
            user.books_set.set(books)
            return Response(status=status.HTTP_202_ACCEPTED)
        
    @action(detail=True, methods=['get', 'patch'], permission_classes=[IsAdminUser])
    def change_exam_noncountable(self, request, pk=None):
        user = self.get_object()

        # اگر ExamAverage وجود نداشت، None برگردانیم یا ایجاد کنیم؟
        exam_avg = getattr(user, 'examaverage', None)
        if exam_avg is None:
            return Response({"detail": "ExamAverage برای این کاربر تعریف نشده است."}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'GET':
            serializer = ExamAverageNCSerializer(exam_avg)
            return Response(serializer.data)

        elif request.method == 'PATCH':
            new_value = request.data.get('exams_noncountable')
            if new_value is not None:
                exam_avg.non_countable_count = new_value
                exam_avg.save(update_fields=['non_countable_count',])
                # اگر متدی برای آپدیت میانگین داری
                if hasattr(exam_avg, 'get_average'):
                    exam_avg.get_average()
            return Response(status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser])
    def classroom_presence_summary(self, request, pk=None):
        user = self.get_object()
    
        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
    
        # Create a worksheet and set the column headers.
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'User')
        worksheet.write(0, 1, 'Classroom Name')
        worksheet.write(0, 2, 'Presence Percentage')
        worksheet.write(0, 3, 'Presence')
    
        # Get the data and write it to the worksheet.
        row = 1
        presences = ClassroomPresence.objects.select_related('classroom', 'classroom_average_reffer__user').filter(classroom_average_reffer__user=user)
        for presence in presences:
            worksheet.write(row, 0, presence.classroom_average_reffer.user.username)
            worksheet.write(row, 1, presence.classroom.ClassroomName)
            worksheet.write(row, 2, presence.classroom_presence_percentage)
            worksheet.write(row, 3, str(presence.classroom_presence))
            row += 1
    
        # Close the workbook and construct the response.
        workbook.close()
        output.seek(0)
        filename = f'{user.username}_classroom_presence_summary.xlsx'
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser])
    def year_records(self, request, pk=None):
        """آرشیو سال‌به‌سال دانش‌آموز (برای نمودار بالای پروفایل)."""
        user = self.get_object()
        records = user.year_records.all().prefetch_related('exam_records', 'assignment_records')
        return Response(StudentYearRecordSerializer(records, many=True).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser], url_path='attendance')
    def attendance(self, request, pk=None):
        """تاریخچه‌ی حضور و غیابِ حضوریِ دانش‌آموز + خلاصه."""
        user = self.get_object()
        records = user.attendance_records.all()
        total = records.count()
        absent = records.filter(present=False).count()
        return Response({
            'total_sessions': total,
            'present_count': total - absent,
            'absent_count': absent,
            'records': AttendanceRecordSerializer(records, many=True).data,
        }, status=status.HTTP_200_OK)

class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all().order_by('-created_at')
    serializer_class = NotificationSerializer
    permission_classes = [IsAdminUser, IsAuthenticated]
    http_method_names = ['get','post','delete','patch']
    def perform_create(self, serializer):
        notif = serializer.save()
        if not notif.is_persistent:
            # تمام کاربران همه گروه‌های انتخاب شده
            users = User.objects.filter(groups__in=notif.groups.all()).distinct().values_list('id', flat=True)
            UserNotification.objects.bulk_create(
                [UserNotification(user_id=u, notification=notif) for u in users]
            )

    @action(detail=True, methods=["get"], url_path="user", permission_classes=[IsAuthenticated])
    def user_notifications(self, request, pk=None):
        username = pk
        if not username:
            return Response({"error": "username required"}, status=status.HTTP_400_BAD_REQUEST)

        # کاربر عادی فقط نوتیف خودش را می‌بیند؛ فقط ادمین/استاف می‌تواند نوتیف دیگران را بخواند
        if not request.user.is_staff and username != request.user.username:
            return Response({"message": "You are not allowed"}, status=status.HTTP_403_FORBIDDEN)

        # نوتیف دائم گروه‌های کاربر
        persistent = Notification.objects.filter(
            groups__user__username=username, is_persistent=True, is_finished=False
        ).values("id", "title", "message","is_persistent")

        # نوتیف یکبارمصرف از UserNotification
        one_time = UserNotification.objects.filter(user__username=username).select_related("notification")
        return Response(list(persistent) + UserNotificationSerializer(one_time, many=True).data , status=status.HTTP_200_OK)

    @action(detail=True, methods=["delete"], url_path="read", permission_classes=[IsAuthenticated])
    def mark_as_read(self, request, pk=None):
        user_notif = get_object_or_404(UserNotification.objects.select_related("notification"), pk=pk)

        # کاربر عادی فقط می‌تواند نوتیف متعلق به خودش را ببندد
        if not request.user.is_staff and user_notif.user_id != request.user.id:
            return Response({"message": "You are not allowed"}, status=status.HTTP_403_FORBIDDEN)

        notif = user_notif.notification
        user_notif.delete()
        if not notif.user_notifications.exists():
            notif.is_finished = True
            notif.save(update_fields=["is_finished"])
        return Response({"status": "Read And Removed"},status=status.HTTP_200_OK)

@extend_schema_view(
    custom=extend_schema(
        request=inline_serializer(name='CustomSmsRequest', fields={
            'group_ids': rfs.ListField(child=rfs.IntegerField(), required=False, help_text='لیست id گروه‌ها'),
            'user_ids': rfs.ListField(child=rfs.IntegerField(), required=False, help_text='لیست id افراد'),
            'message': rfs.CharField(help_text='متن پیام'),
            'target': rfs.ChoiceField(choices=['mother', 'father', 'both'], required=False, help_text='پیش‌فرض mother'),
        }),
        responses=OpenApiTypes.OBJECT),
)
class SMSManagerIndex(viewsets.ViewSet):
    permission_classes = [IsAdminUser, IsAuthenticated]
    # محدودیت نرخ اختصاصی (اسکوپ 'sms') چون ارسال پیامک هزینه‌ی مالی دارد
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = 'sms'

    @action(detail=False, methods=['post'], url_path='custom')
    def custom(self, request):
        """ارسال پیام دلخواه به یک/چند گروه و/یا یک/چند فرد.
        body: {"group_ids":[...], "user_ids":[...], "message":"...", "target":"mother"|"father"|"both"}
        (مقصد پیش‌فرضِ دانش‌آموز را تغییر نمی‌دهد)"""
        logger = logging.getLogger('sms_manager')
        group_ids = request.data.get('group_ids') or []
        user_ids = request.data.get('user_ids') or []
        message = (request.data.get('message') or '').strip()
        target = (request.data.get('target') or 'mother').strip()

        if not message:
            return Response({"error": "message is required"}, status=status.HTTP_400_BAD_REQUEST)
        if not group_ids and not user_ids:
            return Response({"error": "at least one of group_ids or user_ids is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            result = sms_manager.send_custom(group_ids, user_ids, message, target)
            logger.info(f"Custom SMS by {request.user.username} | groups={group_ids} users={user_ids} target={target} | {result}")
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Custom SMS error by {request.user.username}: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#########################   CLASSROOM All Upload Set
class ClassroomsIndex (viewsets.ModelViewSet):
    queryset = Classroom.objects.all().order_by('-classroom_creation_time')
    serializer_class = ClassroomSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]
    search_fields=('ClassroomName','classroom_headline','classroom_groups__name')    
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ClassroomsSerializer
        return ClassroomSerializer
        
    def perform_create(self, serializer):
        instance = serializer.save()
        instance.create_classroom_presence()
        process_content_urls(instance)

    def partial_update(self, request, *args, **kwargs):
        response = super().partial_update(request, *args, **kwargs)
        instance = self.get_object()
        instance.update_classroom_presence()
        process_content_urls(instance)
        return response
#########################    ASSIGNMENT   All Upload Set         
class AssignmentsIndex (viewsets.ModelViewSet):
    queryset = Assignment.objects.all().order_by('-assignment_creation_time')
    serializer_class = AssignmentSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated,IsStaffUser]
    search_fields=('AssignmentName','assignment_headline','assignment_group__name','assignment_available_time_start','assignment_available_time_end')
    # def get_queryset(self):
    #     queryset = Assignment.objects.all()
    #     ordering = self.request.query_params.get('ordering', '-assignment_creation_time')
    #     if ordering is not None:
    #         queryset = queryset.order_by(ordering)
    #     return queryset
        
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return AssignmentSerializerWithGroup
        return super().get_serializer_class()
        
    def perform_create(self, serializer):
        instance = serializer.save()
        # هندل فایل‌های ارسالی
        file_obj = self.request.FILES.get('assignment_file')
        if file_obj:
            # print(instance.AssignmentName)
            # print(instance.assignment_group.name)
            download_url = auto_upload("assignment", instance, file_obj)
            instance.assignment_file = download_url

        file_obj_answer = self.request.FILES.get('assignment_answer_file')
        if file_obj_answer:
            download_url = auto_upload("assignment_answer", instance, file_obj_answer)
            instance.assignment_answer_file = download_url

        instance.save(update_fields=['assignment_file', 'assignment_answer_file'])
        instance.create_assignment_score()

    def partial_update(self, request, *args, **kwargs):
        response = super().partial_update(request, *args, **kwargs)
        instance = self.get_object()

        updated_fields = []
        logger = logging.getLogger('management_logger')
        # ---------------------- فایل‌ها ----------------------
        file_assignment = request.FILES.get('assignment_file')
        file_assignment_answer = request.FILES.get('assignment_answer_file')

        if file_assignment:
            instance.assignment_file = auto_upload('assignment', instance, file_assignment)
            updated_fields.append("assignment_file")

        if file_assignment_answer:
            instance.assignment_answer_file = auto_upload('assignment_answer', instance, file_assignment_answer)
            updated_fields.append("assignment_answer_file")

        # ---------------------- بررسی ریست فقط اگر زمان جدید جلوتر از فعلی باشه ----------------------
        try:
            now = timezone.now()
            reset_triggered = False
            new_end_raw = request.data.get("assignment_available_time_end")

            if instance.assignment_finished and new_end_raw:
                # 1) پارس رشته مثل "2025-08-29 17:32"
                try:
                    new_end_dt = datetime.strptime(new_end_raw, "%Y-%m-%d %H:%M")
                except ValueError:
                    return Response(
                        {"message": "Invalid end time format. Use 'YYYY-MM-DD HH:MM'."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # 2) aware کردن
                if timezone.is_naive(new_end_dt):
                    new_end_dt = timezone.make_aware(new_end_dt, timezone.get_current_timezone())

                # 3) فقط اگر ددلاین قبلی گذشته و جدید جلوتره، ریست کن
                if new_end_dt > now:
                    reset_triggered = True
                    instance.assignment_available_time_end = new_end_dt
                    updated_fields.append("assignment_available_time_end")

            if reset_triggered:
                instance.assignment_finished = False
                logger.info(f"{request.user.username} Reopen Assignment {instance.assignment_id}")
                updated_fields += ["assignment_finished"]
                AssignmentScore.objects.filter(assignment=instance).update(assignment_finished=False)

        except Exception as e:
            logger.error(f"[PATCH ERROR] assignment reset failed: {e}", exc_info=True)

        # ---------------------- ذخیره در صورت نیاز ----------------------
        if updated_fields:
            instance.save(update_fields=list(set(updated_fields)))

        # ---------------------- به‌روزرسانی نمرات ----------------------
        try:
            instance.create_assignment_score()
            instance.update_assignment_score()
        except Exception as e:
            logger.info(f"[PATCH WARNING] assignment score update failed: {str(e)}")
            
        logger.info(f"{request.user.username} Changed Assignment {instance.assignment_id}")
        return response
    
    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser], url_path='sms-recipients')
    def sms_recipients(self, request, pk=None):
        """لیست افراد این تکلیف برای انتخاب در فرانت (نام، نمره، داشتنِ شماره‌ی مادر/پدر)."""
        assignment = get_object_or_404(Assignment, assignment_id=pk)
        return Response(sms_manager.assignment_recipients(assignment), status=status.HTTP_200_OK)

    @extend_schema(
        request=inline_serializer(name='AssignmentSendSmsRequest', fields={
            'user_ids': rfs.JSONField(required=False, help_text='لیست id افراد یا "all" (پیش‌فرض all)'),
            'target': rfs.ChoiceField(choices=['mother', 'father', 'both'], required=False, help_text='پیش‌فرض mother'),
        }),
        responses=OpenApiTypes.OBJECT)
    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser],
            throttle_classes=[SmsRateThrottle], url_path='send-sms')
    def send_sms(self, request, pk=None):
        """ارسال پیامک نمره‌ی این تکلیف.
        body: {"user_ids": [...] یا "all", "target": "mother"|"father"|"both"}"""
        assignment = get_object_or_404(Assignment, assignment_id=pk)
        if not assignment.sms_permission:
            return Response({"error": "هنوز مجوز ارسال پیامک برای این تکلیف صادر نشده (همه تصحیح نشده‌اند)."},
                            status=status.HTTP_400_BAD_REQUEST)
        target = (request.data.get('target') or 'mother').strip()
        user_ids = request.data.get('user_ids', 'all')
        result = sms_manager.send_assignment_scores(assignment, user_ids, target)
        return Response(result, status=status.HTTP_200_OK)
     
class AssignmentScoresIndex (viewsets.ModelViewSet):
    queryset = AssignmentScore.objects.filter(assignment_presence=True).order_by('-updated_file_at')
    serializer_class = AssignmentScoreSerializer
    http_method_names = ['get','delete','patch']   
    permission_classes = [IsAdminUser,IsAuthenticated,IsStaffUser]
    search_fields=('assignment_average_reffer__user__first_name','assignment_average_reffer__user__last_name','assignment__AssignmentName','assignment_presence','assignment_finished','assignment_marked','assignment_marked_by')
    
    def get_queryset(self):
        # برای صفحه‌ی index همان قبلی بماند، اما اکشن manual به همه دسترسی داشته باشد
        if getattr(self, 'action', None) == 'manual_assignmentscore':
            return AssignmentScore.objects.all().order_by('-updated_file_at')
        return super().get_queryset()

    def get_serializer_class(self):
        if self.action == 'list':
            return AssignmentScoresSerializer
        return AssignmentScoreSerializer   
    
    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        logger = logging.getLogger('admins_manager')
        instance.assignment_marked = True
        instance.assignment_permission = False
        instance.assignment_marked_by = f"{request.user.first_name} {request.user.last_name}"

        # هندل فایل تصحیح‌شده
        teacher_file = request.FILES.get('assignment_teacher_file')
        if teacher_file:
            instance.assignment_teacher_file = auto_upload("assignment_teacher", instance, teacher_file)
            # ذخیره در مدل
            instance.save(update_fields=['assignment_teacher_file', 'assignment_marked', 'assignment_permission', 'assignment_marked_by'])
        else:
            # اگر فایل آپلود نشده بود، فقط فلگ‌ها رو آپدیت کن
            instance.save(update_fields=['assignment_marked', 'assignment_permission', 'assignment_marked_by'])

        # ادامه پروسه اصلی
        response = super().partial_update(request, *args, **kwargs)
        instance.refresh_from_db()
        instance.get_score()
        # اگر همه‌ی نمرات این تکلیف تصحیح شدند، مجوز ارسال پیامک صادر می‌شود
        instance.assignment.update_sms_permission()
        logger.info(f"Assignment {instance.pk} Marked By:{request.user.username}")
        return response
    
    # @swagger_auto_schema(auto_schema=None)
    @action(detail=True,methods=['patch'],permission_classes=[IsAdminUser],parser_classes=[parsers.MultiPartParser, parsers.FormParser])
    def manual_assignmentscore(self, request, pk=None):
        try:
            instance: AssignmentScore = self.get_object()
            admin_user = request.user.get_username()
            logger = logging.getLogger('management_logger')
            logger.info(f"Manual upload sending by '{admin_user}' - {instance.pk}")

            # دریافت فایل
            file_obj = request.FILES.get('file') or request.data.get('file')
            if not file_obj or not hasattr(file_obj, 'name'):
                return Response({"message": "No file provided. Use form field 'file'."},
                                status=status.HTTP_400_BAD_REQUEST)

            # توجه: خارج از محدودیت — چک نوع/سایز اجباری نیست.
            # اگر خواستی فقط PDF بپذیری، این دو خط را uncomment کن:
            if not file_obj or not hasattr(file_obj, 'size'):
                logger.error("* No file was uploaded or the uploaded object is not a file.")
                return response.Response({"message": "No file was uploaded or invalid file type"}, status=status.HTTP_400_BAD_REQUEST)
            # logger.info(f"Received file: {file_obj.name} with size: {file_obj.size} bytes")            
            if file_obj.size > 10 * 1024 * 1024:
                logger.error("* Upload attempt failed - file size exceeds limit.")
                return response.Response({"message": "File too large"}, status=status.HTTP_400_BAD_REQUEST)

            if not file_obj.name.lower().endswith('.pdf'):
                logger.error("* Upload attempt failed - incorrect file type.")
                return response.Response({"message": "Invalid file type"}, status=status.HTTP_400_BAD_REQUEST)

            # ذخیره فایل با همان util پروژه
            stored_path = auto_upload("assignment_student", instance, file_obj)

            # به‌روزرسانی وضعیت‌ها
            instance.assignment_student_file = stored_path
            instance.assignment_presence = True
            instance.assignment_marked = False
            instance.assignment_marked_by = f"{request.user.first_name} {request.user.last_name}"
            instance.updated_file_at = timezone.now()
            instance.save(update_fields=[
                'assignment_student_file',
                'assignment_presence',
                'assignment_marked',
                'assignment_marked_by',
                'updated_file_at'
            ])

            logger.info(f"Manual upload OK by '{admin_user}' - {instance.pk}")
            return Response({
                "message": "File uploaded successfully by admin.",
                "file_url": instance.assignment_student_file
            }, status=status.HTTP_200_OK)

        except AssignmentScore.DoesNotExist:
            return Response({"message": "AssignmentScore not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("Manual upload failed")
            return Response({"message": f"An error occurred: {str(e)}"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                     
    # @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    # def identify_orphaned_assignments(self, request):
    #     scripts.identify_orphaned_files()
    #     return Response({"message": "Orphaned assignments identified and logged to 'orphaned_files.txt'."}, status=status.HTTP_200_OK)

    # @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    # def delete_orphaned_assignments(self, request):
    #     scripts.delete_orphaned_files()
    #     return Response({"message": "Orphaned assignments deleted and logged."}, status=status.HTTP_200_OK)
    # @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    # def debug_orphaned_assignments(self, request):
    #     scripts.debug_list_assignments()
    #     return Response({"message": "Debug scan completed and logged."}, status=status.HTTP_200_OK)

#########################   EXAMS Upload Done
class ExamsIndex (viewsets.ModelViewSet):
    queryset = Exam.objects.all().order_by('-exam_available_time_start')
    serializer_class = ExamSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]
    search_fields=('ExamName','exam_available_time_start','exam_group__name')
    # def list(self, request, pk=None):
    #     ordering = request.query_params.get('ordering', '-exam_available_time_start',)
    #     queryset = Exam.objects.all().order_by(ordering)
    #     serializer = ExamSerializerWithGroup(queryset, many=True)
    #     return Response(serializer.data)
    
    # def retrieve(self, request, pk=None):
    #     # ordering = request.query_params.get('ordering', '-exam_available_time_start', '-exam_available_time_start','exam_creation_time')
    #     queryset = Exam.objects.all()#.order_by(ordering)
    #     exam = get_object_or_404(queryset, pk=pk)
    #     serializer = ExamSerializerWithQuestion(exam)
    #     return Response(serializer.data) 
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ExamSerializerWithGroup
        if self.action == 'retrieve':
            return ExamSerializerWithQuestion    
        return ExamSerializer
            
    def perform_create(self, serializer):
        instance = serializer.save()
        # هندل آپلود فایل آزمون
        exam_description_file = self.request.FILES.get('exam_description')
        exam_answer_file = self.request.FILES.get('exam_answer_file')

        if exam_description_file:
            uploaded_path = auto_upload("exam_description",instance,exam_description_file)
            instance.exam_description = uploaded_path

        if exam_answer_file:
            uploaded_path = auto_upload("exam_answer",instance,exam_answer_file)
            instance.exam_answer_file = uploaded_path
        instance.save()
        # سایر عملیات
        instance.merge_question_answer_images()
        instance.set_exam_time()
        instance.create_exam_score()
        # eta = instance.exam_available_time_end + timedelta(seconds=1)
        # finish_exam.apply_async((instance.exam_id,), eta=eta)

    def partial_update(self, request, *args, **kwargs):
        response = super().partial_update(request, *args, **kwargs)
        instance = self.get_object()

        # آپلود در صورت وجود فایل جدید
        exam_description_file = request.FILES.get('exam_description')
        exam_answer_file = request.FILES.get('exam_answer_file')

        if exam_description_file:
            uploaded_path = auto_upload("exam_description", instance, exam_description_file)
            instance.exam_description = uploaded_path

        if exam_answer_file:
            uploaded_path = auto_upload("exam_answer", instance, exam_answer_file)
            instance.exam_answer_file = uploaded_path
            
        instance.save()
        instance.merge_question_answer_images()
        instance.set_exam_time()
        instance.create_exam_score()
        instance.update_exam_score()
        # instance.finish_exam()
        return response
    
    @action(detail=True, methods=['get'], permission_classes=[IsAdminUser], url_path='sms-recipients')
    def sms_recipients(self, request, pk=None):
        """لیست افراد این آزمون برای انتخاب در فرانت (نام، نمره، داشتنِ شماره‌ی مادر/پدر)."""
        exam = get_object_or_404(Exam, exam_id=pk)
        return Response(sms_manager.exam_recipients(exam), status=status.HTTP_200_OK)

    @extend_schema(
        request=inline_serializer(name='ExamSendSmsRequest', fields={
            'user_ids': rfs.JSONField(required=False, help_text='لیست id افراد یا "all" (پیش‌فرض all)'),
            'target': rfs.ChoiceField(choices=['mother', 'father', 'both'], required=False, help_text='پیش‌فرض mother'),
        }),
        responses=OpenApiTypes.OBJECT)
    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser],
            throttle_classes=[SmsRateThrottle], url_path='send-sms')
    def send_sms(self, request, pk=None):
        """ارسال پیامک نمره‌ی این آزمون (شامل آزمون حضوری/آفلاین).
        body: {"user_ids": [...] یا "all", "target": "mother"|"father"|"both"}"""
        exam = get_object_or_404(Exam, exam_id=pk)
        if not exam.sms_permission:
            return Response({"error": "هنوز مجوز ارسال پیامک برای این آزمون صادر نشده."},
                            status=status.HTTP_400_BAD_REQUEST)
        target = (request.data.get('target') or 'mother').strip()
        user_ids = request.data.get('user_ids', 'all')
        result = sms_manager.send_exam_scores(exam, user_ids, target)
        return Response(result, status=status.HTTP_200_OK)

class ExamScoresIndex (viewsets.ModelViewSet):
    # queryset = ExamScore.objects.all()
    serializer_class = ExamScoresSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]
    
    def get_queryset(self):
        queryset = ExamScore.objects.all()
        ordering = self.request.query_params.get('ordering', '-updated_at')
        if ordering is not None:
            queryset = queryset.order_by(ordering)
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ExamScoresSerializerWithExamName
        return ExamScoresSerializer
        
    @action(detail=True, methods=['get',], permission_classes=[IsAdminUser])
    def user_scores_list(self, request, pk=None):
        exam_scores = ExamScore.objects.filter(exam_average_reffer__user__pk=pk)
        # serializer = self.get_serializer(exam_scores, many=True)
        serializer = ExamScoresSerializerWithExamName(exam_scores, many=True)
        return Response(serializer.data)
        
@extend_schema_view(
    retrieve=extend_schema(parameters=[OpenApiParameter('id', OpenApiTypes.UUID, OpenApiParameter.PATH)]),
    partial_update=extend_schema(parameters=[OpenApiParameter('id', OpenApiTypes.UUID, OpenApiParameter.PATH)]),
    destroy=extend_schema(parameters=[OpenApiParameter('id', OpenApiTypes.UUID, OpenApiParameter.PATH)]),
)
class QuestionsIndex(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    http_method_names = ['get', 'post', 'delete', 'patch']
    permission_classes = [IsAdminUser, IsAuthenticated]
    search_fields = ('question_headline', 'question_category', 'question_grade', 'question_book')

    def get_queryset(self):
        queryset = Question.objects.all()
        ordering = self.request.query_params.get('ordering', '-question_creation_time')
        if ordering is not None:
            queryset = queryset.order_by(ordering)
        return queryset
    
    def perform_create(self, serializer):
        instance = serializer.save()

        # بررسی فایل‌های ارسالی برای سوال
        question_img_file = self.request.FILES.get('question_img')
        question_answer_img_file = self.request.FILES.get('question_answer_img')

        if question_img_file:
            uploaded_path = auto_upload("question", instance, question_img_file)
            instance.question_img = uploaded_path

        if question_answer_img_file:
            uploaded_path = auto_upload("question_answer", instance, question_answer_img_file)
            instance.question_answer_img = uploaded_path
            
        instance.save()
        
    def partial_update(self, request, *args, **kwargs):
        response = super().partial_update(request, *args, **kwargs)
        instance = self.get_object()

        # آپلود در صورت وجود فایل جدید
        question_img_file = request.FILES.get('question_img')
        question_answer_img_file = request.FILES.get('question_answer_img')

        if question_img_file:
            uploaded_path = auto_upload("question", instance, question_img_file)
            instance.question_img = uploaded_path

        if question_answer_img_file:
            uploaded_path = auto_upload("question_answer", instance, question_answer_img_file)
            instance.question_answer_img = uploaded_path

        instance.save()
        return response    

##################  MONEY  
class SignedCheckIndex (viewsets.ModelViewSet):
    queryset = SignedCheck.objects.all()
    serializer_class = SignedCheckSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]

class BooksIndex (viewsets.ModelViewSet):
    queryset = Books.objects.all()
    serializer_class = BooksSerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]

class DirectMoneyIndex (viewsets.ModelViewSet):
    queryset = DirectMoney.objects.all()
    serializer_class = DirectMoneySerializer
    http_method_names = ['get','post','delete','patch']
    permission_classes = [IsAdminUser,IsAuthenticated]

############################# Uploads
  
###########Upload working stable version
@extend_schema_view(post=extend_schema(
    request=inline_serializer(name='UploadExcelRequest', fields={
        'file': rfs.FileField(help_text='فایل اکسل دانش‌آموزان'),
    }),
    responses=OpenApiTypes.OBJECT))
class UploadExcelView(views.APIView):
    http_method_names = ['post']
    permission_classes = [IsAdminUser, IsAuthenticated]

    def post(self, request):
        file = request.FILES['file']
        data = pd.read_excel(file, dtype={
            'کد ملی': str,
            'شماره دانش آموز': str,
            'شماره پدر': str,
            'شماره مادر': str,
            'شماره منزل': str
        })

        is_admin_file = 'admin' in data.columns and str(data.loc[0, 'admin']).strip() == 'ntorabi'

        # ساخت اکانت staff فقط توسط superuser مجاز است (جلوگیری از ارتقای دسترسی)
        if is_admin_file and not request.user.is_superuser:
            return Response({"error": "Only a superuser can upload an admin (staff-creating) file"},
                            status=status.HTTP_403_FORBIDDEN)

        field_mapping = {
            'father_name': 'نام پدر',
            'phone_number': 'شماره دانش آموز',
            'father_number': 'شماره پدر',
            'mother_number': 'شماره مادر',
            'home_number': 'شماره منزل',
            'address': 'آدرس',
            'student_school': 'مدرسه',
            'student_type': 'رشته',
            'student_gender': 'جنسیت',
            'student_grade': 'مقطع'
        }

        for _, row in data.iterrows():

            # حالت فایل admin: فقط یوزر بساز، بقیه رد شو
            if is_admin_file:
                if pd.isna(row.get('نام کاربری')) or pd.isna(row.get('رمز عبور')):
                    continue
                username = str(row['نام کاربری']).strip()
                password = str(row['رمز عبور']).strip()
                first_name = str(row['نام']).strip() if not pd.isna(row.get('نام')) else ""
                last_name = str(row['نام خانوادگی']).strip() if not pd.isna(row.get('نام خانوادگی')) else ""

                user, created = User.objects.get_or_create(username=username)
                if created:
                    user.set_password(password)
                user.first_name = first_name
                user.last_name = last_name
                user.is_staff = True
                if 'وضعیت' in row and not pd.isna(row['وضعیت']):
                    status_value = str(row['وضعیت']).strip().lower()
                    if status_value == "active":
                        user.is_active = True
                    elif status_value == "deactive":
                        user.is_active = False
                user.save()
                continue
            # حالت فایل عادی: روند کامل
            if pd.isna(row.get('کد ملی')):
                continue
            username = str(row['کد ملی']).strip()
            password = username[-4:]

            user, created = User.objects.get_or_create(username=username)
            if created:
                user.set_password(password)
            user.first_name = row.get('نام', "")
            user.last_name = row.get('نام خانوادگی', "")
            # دانش‌آموزِ حاضر در اکسلِ سال جدید فعال می‌شود (برگشتی‌ها بعد از بایگانی دوباره فعال)
            user.is_active = True
            user.save()

            group_name = row['گروه']
            group, created = Group.objects.get_or_create(name=group_name)
            if created:
                group.group_time = row['ساعت کلاس']
                group.group_day = row['روز کلاس']
                group.group_grade = row['مقطع']
                group.group_gender = row['جنسیت']
                group.save()
            user.groups.clear()
            user.groups.add(group)

            student_user, created = StudentUser.objects.get_or_create(student_user=user)
            for field, column in field_mapping.items():
                if column in row and not pd.isna(row[column]):
                    setattr(student_user, field, row[column])
            student_user.student_status = 'در حال تحصیل'
            student_user.registration_date = date.today()
            student_user.save()
            student_user.create_averages()

            if not pd.isna(row.get('مبلغ چک')):
                try:
                    check_amount = str(int(float(row['مبلغ چک'])))
                except:
                    check_amount = "0"
                check_number = str(row.get('شماره چک')) if not pd.isna(row.get('شماره چک')) else "-"
                bank = str(row.get('نام بانک')) if not pd.isna(row.get('نام بانک')) else "-"
                if not SignedCheck.objects.filter(student=user, amout=check_amount, check_number=check_number, bank=bank).exists():
                    try:
                        y, m, d = map(int, row['تاریخ چک'].split("/"))
                        check_date = jdatetime.date(y, m, d).togregorian().strftime("%Y-%m-%d")
                    except:
                        check_date = "-"
                    SignedCheck.objects.create(
                        student=user,
                        amout=check_amount,
                        check_date=check_date,
                        check_number=check_number,
                        bank=bank,
                        description=str(row.get('توضیحات چک')) if not pd.isna(row.get('توضیحات چک')) else "-"
                    )

            if not pd.isna(row.get('مبلغ نقدی')):
                try:
                    direct_amount = str(int(float(row['مبلغ نقدی'])))
                except:
                    direct_amount = "0"
                following = str(row.get('شماره پیگیری')) if not pd.isna(row.get('شماره پیگیری')) else "-"
                card = str(row.get('چهار رقم کارت')) if not pd.isna(row.get('چهار رقم کارت')) else "-"
                if not DirectMoney.objects.filter(student=user, amout=direct_amount, following_number=following, card_number=card).exists():
                    try:
                        y, m, d = map(int, row['تاریخ پرداخت'].split("/"))
                        payment_date = jdatetime.date(y, m, d).togregorian().strftime("%Y-%m-%d")
                    except:
                        payment_date = "-"
                    DirectMoney.objects.create(
                        student=user,
                        amout=direct_amount,
                        payment_date=payment_date,
                        refrence_number=str(row.get('شماره مرجع')) if not pd.isna(row.get('شماره مرجع')) else "-",
                        following_number=following,
                        card_number=card,
                        payment_method=str(row.get('انتقال / POS')) if not pd.isna(row.get('انتقال / POS')) else "-",
                        bank=str(row.get('بانک عامل')) if not pd.isna(row.get('بانک عامل')) else "-",
                        description=str(row.get('توضیحات نقدی')) if not pd.isna(row.get('توضیحات نقدی')) else "-"
                    )

        return Response(status=status.HTTP_201_CREATED)


# ───────────────── همگام‌سازی اکسلِ امتحان حضوری و حضور و غیاب ─────────────────
def _parse_jalali_date(s):
    """'1403/07/15' → datetime آگاه از تایم‌زون (نیمه‌شب تهران)؛ None اگر نامعتبر."""
    try:
        y, m, d = map(int, str(s).strip().split("/"))
        g = jdatetime.date(y, m, d).togregorian()
        dt = datetime(g.year, g.month, g.day)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    except Exception:
        return None


def _is_absent_value(v):
    """تفسیر مقدار وضعیت/حاضر در اکسل → True یعنی غایب."""
    return str(v).strip() in ('غایب', 'absent', 'Absent', '0', 'false', 'False', 'خیر')


@extend_schema_view(post=extend_schema(
    request=inline_serializer(name='UploadOfflineExamRequest', fields={
        'file': rfs.FileField(help_text='اکسل نتایج (ستون‌ها: کد ملی، درصد، حاضر?)'),
        'exam_name': rfs.CharField(),
        'group': rfs.CharField(help_text='نام گروه'),
        'date': rfs.CharField(help_text="تاریخ جلالی مثل 1403/07/15"),
        'headline': rfs.CharField(required=False),
    }),
    responses=OpenApiTypes.OBJECT))
class UploadOfflineExamView(views.APIView):
    """
    آپلود نتایج یک امتحان حضوری (آفلاین).
    فیلدهای فرم: exam_name، group (نام گروه)، date ('1403/07/15')، headline (اختیاری)
    ستون‌های اکسل: 'کد ملی' + 'درصد'  (+ اختیاری 'حاضر')
    یک Exam سبک می‌سازد و برای هر دانش‌آموز ExamScoreOffline ثبت می‌کند که در میانگین امتحانات لحاظ می‌شود.
    اجرای مجدد با همان exam_name+group، نمرات قبلی همان امتحان را بازنویسی می‌کند (idempotent).
    """
    parser_classes = (parsers.MultiPartParser, parsers.FormParser)
    permission_classes = [IsAdminUser, IsAuthenticated]

    def post(self, request):
        logger = logging.getLogger('management_logger')
        file       = request.FILES.get('file')
        exam_name  = (request.data.get('exam_name') or '').strip()
        group_name = (request.data.get('group') or '').strip()
        headline   = (request.data.get('headline') or '').strip()
        date_dt    = _parse_jalali_date(request.data.get('date'))

        if not file or not exam_name or not group_name:
            return Response({"error": "file, exam_name و group لازم‌اند"}, status=status.HTTP_400_BAD_REQUEST)

        group = Group.objects.filter(name=group_name).first()
        if not group:
            return Response({"error": f"گروه '{group_name}' یافت نشد"}, status=status.HTTP_404_NOT_FOUND)

        try:
            data = pd.read_excel(file, dtype={'کد ملی': str})
        except Exception as e:
            return Response({"error": f"خطا در خواندن اکسل: {e}"}, status=status.HTTP_400_BAD_REQUEST)
        if 'کد ملی' not in data.columns or 'درصد' not in data.columns:
            return Response({"error": "ستون‌های 'کد ملی' و 'درصد' لازم‌اند"}, status=status.HTTP_400_BAD_REQUEST)

        skipped, applied, affected = [], 0, set()
        with transaction.atomic():
            exam, _ = Exam.objects.get_or_create(
                ExamName=exam_name, exam_group=group,
                defaults=dict(exam_headline=(headline or exam_name), exam_finished=True, exam_permission=False),
            )
            if date_dt:
                exam.exam_available_time_start = date_dt
                exam.exam_available_time_end = date_dt
            exam.exam_headline = headline or exam.exam_headline
            exam.exam_finished = True
            exam.exam_permission = False
            exam.sms_permission = True   # امتحان حضوری ثبت شد → مجوز ارسال پیامک
            exam.save()
            ExamScoreOffline.objects.filter(exam=exam).delete()  # بازنویسی روی اجرای مجدد

            for _, row in data.iterrows():
                if pd.isna(row.get('کد ملی')):
                    continue
                username = str(row['کد ملی']).strip()
                user = User.objects.filter(username=username).first()
                if not user:
                    skipped.append(username); continue
                exam_avg, _ = ExamAverage.objects.get_or_create(user=user)
                present = not ('حاضر' in row and not pd.isna(row['حاضر']) and _is_absent_value(row['حاضر']))
                try:
                    score = round(float(row['درصد']), 2) if present and not pd.isna(row.get('درصد')) else 0
                except Exception:
                    score = 0
                ExamScoreOffline.objects.create(
                    exam=exam, exam_average_reffer=exam_avg,
                    score=score, exam_peresence=present, countable=True,
                )
                affected.add(exam_avg.id); applied += 1

            for avg in ExamAverage.objects.filter(id__in=affected):
                avg.get_average()

        logger.info(f"{request.user.username} offline-exam '{exam_name}' group={group_name}: {applied} scores, {len(skipped)} skipped")
        return Response({"message": "نتایج امتحان حضوری ثبت شد", "exam_id": str(exam.exam_id),
                         "applied": applied, "skipped": skipped}, status=status.HTTP_200_OK)


@extend_schema_view(post=extend_schema(
    request=inline_serializer(name='UploadAttendanceRequest', fields={
        'file': rfs.FileField(help_text='اکسل حضور و غیاب (ستون‌ها: کد ملی، وضعیت?)'),
        'group': rfs.CharField(help_text='نام گروه'),
        'session_title': rfs.CharField(),
        'date': rfs.CharField(help_text="تاریخ جلالی مثل 1403/07/15"),
    }),
    responses=OpenApiTypes.OBJECT))
class UploadAttendanceView(views.APIView):
    """
    آپلود حضور و غیاب یک جلسه‌ی حضوری.
    فیلدهای فرم: group (نام گروه)، date ('1403/07/15')، session_title
    ستون‌های اکسل: 'کد ملی'  (+ اختیاری 'وضعیت' = حاضر/غایب؛ پیش‌فرض حاضر)
    برای هر دانش‌آموز یک AttendanceRecord می‌سازد و شمارنده‌ی غیبت (ClassroomAverage) را از روی رکوردها بازمحاسبه می‌کند.
    اجرای مجدد همان جلسه (تاریخ+موضوع) به‌جای تکرار، رکورد را به‌روزرسانی می‌کند (idempotent).
    """
    parser_classes = (parsers.MultiPartParser, parsers.FormParser)
    permission_classes = [IsAdminUser, IsAuthenticated]

    def post(self, request):
        logger = logging.getLogger('management_logger')
        file          = request.FILES.get('file')
        group_name    = (request.data.get('group') or '').strip()
        session_title = (request.data.get('session_title') or '').strip()
        date_dt       = _parse_jalali_date(request.data.get('date'))

        if not file or not session_title or date_dt is None:
            return Response({"error": "file, session_title و date ('1403/07/15') لازم‌اند"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            data = pd.read_excel(file, dtype={'کد ملی': str})
        except Exception as e:
            return Response({"error": f"خطا در خواندن اکسل: {e}"}, status=status.HTTP_400_BAD_REQUEST)
        if 'کد ملی' not in data.columns:
            return Response({"error": "ستون 'کد ملی' لازم است"}, status=status.HTTP_400_BAD_REQUEST)

        session_date = date_dt.date()
        skipped, applied, affected = [], 0, set()
        with transaction.atomic():
            for _, row in data.iterrows():
                if pd.isna(row.get('کد ملی')):
                    continue
                username = str(row['کد ملی']).strip()
                user = User.objects.filter(username=username).first()
                if not user:
                    skipped.append(username); continue
                present = not ('وضعیت' in row and not pd.isna(row['وضعیت']) and _is_absent_value(row['وضعیت']))
                AttendanceRecord.objects.update_or_create(
                    student=user, date=session_date, session_title=session_title,
                    defaults=dict(group_name=(group_name or None), present=present),
                )
                affected.add(user.id); applied += 1

            # شمارنده‌ی غیبت را از روی رکوردها بازمحاسبه کن (idempotent)
            for user in User.objects.filter(id__in=affected):
                class_avg, _ = ClassroomAverage.objects.get_or_create(user=user)
                class_avg.absence_count = AttendanceRecord.objects.filter(student=user, present=False).count()
                class_avg.save(update_fields=['absence_count'])

        logger.info(f"{request.user.username} attendance '{session_title}' {session_date} group={group_name}: {applied} rows, {len(skipped)} skipped")
        return Response({"message": "حضور و غیاب ثبت شد", "session_title": session_title,
                         "applied": applied, "skipped": skipped}, status=status.HTTP_200_OK)
    
    
    
    
    
